from fastapi import FastAPI, APIRouter, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from fastapi.responses import FileResponse
from fpdf import FPDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from datetime import date, datetime
import logging

router = APIRouter(prefix="/hdv", tags=["add"])


# Got to fix this model to be on the hdv_sb_models.py
class FormData(BaseModel):
    departamento: str
    municipio: str
    entidad: str
    correo: str
    direccion: str
    telefono: str

    equipo: str
    marca: str
    modelo: str
    serie: str
    activoFijo: str
    registroSanitario: str
    ubicacion: str
    proveedor: str
    img: str

    AdquisitionWay: str
    yearOfFabrication: date
    boughtDate: date
    installationDate: date
    warrantyEnd: date
    fabricante: str

    tension: str
    potencia: str
    presion: str
    corriente: str
    frecuencia: str
    rangoTemperatura: str
    peso: str
    velocidad: str
    tecnologiaPredominante: str

    rangoVoltaje: str
    rangoPresion: str
    rangoHumedad: str
    rangoCorriente: str
    rangoFrecuencia: str

    diagnostico: str
    rehabilitacion: str
    prevencion: str
    tratamientoVida: str
    analisisLaboratorio: str

    clasificacion: str
    periodicidad: str
    calibracion: str


# Route for filling out the excel template with the user's inputs
@router.post("/fill_excel")
async def fill_excel(data: FormData):
    print(type(data.yearOfFabrication))
    try:
        # Cargar el archivo Excel hoja_De_vida.xlsx
        workbook = load_workbook("hdv.xlsx")
        sheet = workbook.active

        # Llenar los campos del archivo Excel con los datos recibidos tipo AllFormData
        # provenientes de Next JS:
        # I. Ubicacion geográfica
        sheet["C6"] = data.departamento
        sheet["C7"] = data.municipio
        sheet["C8"] = data.entidad
        sheet["C9"] = data.correo
        sheet["C10"] = data.direccion
        sheet["C11"] = data.telefono
        sheet["C13"] = data.equipo
        sheet["C14"] = data.marca
        sheet["C15"] = data.modelo
        sheet["C16"] = data.serie
        sheet["C17"] = data.activoFijo
        sheet["C18"] = data.registroSanitario
        sheet["C19"] = data.ubicacion
        sheet["C20"] = data.proveedor
        # Still needing the upload_image endpoint to create a local path for them
        # img_url = data.img
        # img = Image(img_url)
        # img.width = 1000  # Set the width of the image
        # img.height = 100  # Set the height of the image

        # # Add the image to the worksheet
        # img_anchor = "H5"  # The top-left cell for the image
        # sheet.add_image(img, img_anchor)

        # # Merge cells for the combined format
        # start_cell = "H5"
        # end_cell = "J11"
        # sheet.merge_cells(f"{start_cell}:{end_cell}")

        # sheet["G13"] = data.AdquisitionWay
        # date_object = datetime()  # Replace with your datetime object

        # Extract year, month, and day
        year = data.yearOfFabrication.year
        month = data.yearOfFabrication.month
        day = data.yearOfFabrication.day
        sheet["H16"] = day
        sheet["I16"] = month
        sheet["J16"] = year
        year_boughtDate = data.boughtDate.year
        month_boughtDate = data.boughtDate.month
        day_boughtDate = data.boughtDate.day
        sheet["H17"] = day_boughtDate
        sheet["I17"] = month_boughtDate
        sheet["J17"] = year_boughtDate
        year_installationDate = data.installationDate.year
        month_installationDate = data.installationDate.month
        day_installationDate = data.installationDate.day
        sheet["H18"] = day_installationDate
        sheet["I18"] = month_installationDate
        sheet["J18"] = year_installationDate
        year_warrantyEnd = data.warrantyEnd.year
        month_warrantyEnd = data.warrantyEnd.month
        day_warrantyEnd = data.warrantyEnd.day
        sheet["H19"] = day_warrantyEnd
        sheet["I19"] = month_warrantyEnd
        sheet["J19"] = year_warrantyEnd

        # sheet["G17"] = data.warrantyEnd
        sheet["F20"] = data.fabricante

        sheet["C23"] = data.tension
        sheet["C24"] = data.potencia
        sheet["C25"] = data.presion
        sheet["G23"] = data.corriente
        sheet["G24"] = data.frecuencia
        sheet["G25"] = data.rangoTemperatura
        sheet["J23"] = data.peso
        sheet["J24"] = data.velocidad
        sheet["J25"] = data.tecnologiaPredominante

        sheet["C27"] = data.rangoVoltaje
        sheet["C28"] = data.rangoPresion
        sheet["C29"] = data.rangoHumedad
        sheet["C30"] = data.rangoCorriente
        sheet["C31"] = data.rangoFrecuencia

        sheet["H27"] = data.diagnostico
        sheet["H28"] = data.rehabilitacion
        sheet["H29"] = data.prevencion
        sheet["H30"] = data.tratamientoVida
        sheet["H31"] = data.analisisLaboratorio

        sheet["A34"] = data.clasificacion
        sheet["E34"] = data.periodicidad
        sheet["H34"] = data.calibracion

        # Llenar los otros campos del formulario
        # II. ...
        # Guardar el archivo Excel modificado en un directorio temporal
        temp_excel_file_path = "hdv_temp.xlsx"
        workbook.save(temp_excel_file_path)
        return FileResponse(temp_excel_file_path, filename="hdv.xlsx")
    except FileNotFoundError as file_error:
        error_message = f"Error al abrir el archivo de plantilla: {str(file_error)}"
        logging.error(error_message)
        raise HTTPException(status_code=500, detail=error_message)
    except Exception as e:
        # Handle other exceptions and provide a generic error message
        error_message = f"Error al procesar los datos: {str(e)}"
        logging.error(error_message)
        raise HTTPException(status_code=500, detail="Ocurrió un error en el servidor.")
